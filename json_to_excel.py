import json
import openpyxl
import pandas as pd
xlpath = '../EXCELJSON/datafile2.xlsx'
jsonpath = './jsonfile.json'


def json_to_excel():
    workbook = openpyxl.load_workbook(xlpath)
    sheet = workbook['Sheet1']
    with open (jsonpath) as f:
        jsondata = json.load(f)

    for i in range(len(jsondata)):
        item = jsondata[i]
        print(item)
        if i == 0:
            keys = list(item.keys())
            print(keys)
            for k in range(len(keys)):
                sheet.cell(row=(i+1), column=(k+1)).value = keys[k]
        for j in range(len(keys)):
            sheet.cell(row=(i+2), column=(j+1)).value = item[keys[j]]
    workbook.save(xlpath)



json_to_excel()