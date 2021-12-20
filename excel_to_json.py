import openpyxl
import json

xlpath = '../EXCELJSON/datafile.xlsx'
jsonpath = '../jsonfile.json'

ls=[]
def excel_to_json():
    workbook = openpyxl.load_workbook(xlpath)
    sheet = workbook['datainfo']

    for i in range(2, sheet.max_row+1):
        dic = {}
        dic2={}
        dic['id'] = sheet.cell(row=i, column=1).value
        dic['name'] = sheet.cell(row=i, column=2).value
        dic['value'] = sheet.cell(row=i, column=3).value

        ls.append(dic)
    print(ls)

    j = json.dumps(ls)
    with open('jsonfile.json', 'w') as f:
        f.write(j)


excel_to_json()